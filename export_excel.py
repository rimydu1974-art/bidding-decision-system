# -*- coding: utf-8 -*-
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

# 读取AI分析结果
with open(r'C:\Users\ips\Desktop\测试AI员工能力\重做投标网站\杭州非遗项目-评估结果.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = '投标决策评估表'

# 设置列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 20

# 表头样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 写入表头
headers = ['类别', '字段名称', '项目数据', '来源/备注', '风险等级', '是否影响废标', '是否影响得分', '建议动作']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

row = 2

# 辅助函数
def add_row(category, fieldName, projectData, reference='', riskLevel='', affectsVoid='', affectsScore='', suggestion=''):
    global row
    ws.cell(row=row, column=1, value=category)
    ws.cell(row=row, column=2, value=fieldName)
    ws.cell(row=row, column=3, value=projectData)
    ws.cell(row=row, column=4, value=reference)
    ws.cell(row=row, column=5, value=riskLevel)
    ws.cell(row=row, column=6, value=affectsVoid)
    ws.cell(row=row, column=7, value=affectsScore)
    ws.cell(row=row, column=8, value=suggestion)
    row += 1

# ==================== 第1类：基本信息 ====================
basicInfo = data.get('basicInfo', {})
add_row('第1类：基本信息', '项目名称', basicInfo.get('projectName', '-'))
add_row('第1类：基本信息', '项目编号', basicInfo.get('projectCode', '-'))
add_row('第1类：基本信息', '招标企业/采购人', basicInfo.get('tenderer', '-'))
add_row('第1类：基本信息', '联系人', basicInfo.get('contactPerson', '-'))
add_row('第1类：基本信息', '联系电话', basicInfo.get('contactPhone', '-'))
add_row('第1类：基本信息', '代理机构', basicInfo.get('agency', '-'))
add_row('第1类：基本信息', '信息来源', basicInfo.get('informationSource', '-'))
add_row('第1类：基本信息', 'CA需求', basicInfo.get('caRequirement', '-'))
add_row('第1类：基本信息', '开标方式', basicInfo.get('bidOpeningMethod', '-'))
add_row('第1类：基本信息', '开标地点', basicInfo.get('bidOpeningLocation', '-'))
add_row('第1类：基本信息', '报名方式', basicInfo.get('registrationMethod', '-'))
add_row('第1类：基本信息', '项目地点', basicInfo.get('location', '-'))

# ==================== 第2类：财务信息 ====================
financialInfo = data.get('financialInfo', {})
add_row('第2类：财务信息', '资金来源', financialInfo.get('fundingSource', '-'))
add_row('第2类：财务信息', '预算金额', f"{financialInfo.get('budget', 0)}元")
add_row('第2类：财务信息', '最高限价', f"{financialInfo.get('maxPrice', 0)}元")
add_row('第2类：财务信息', '预先投资金额', financialInfo.get('preInvestment', '-'))
add_row('第2类：财务信息', '付款方式', financialInfo.get('paymentMethod', '-'))
add_row('第2类：财务信息', '标书费', financialInfo.get('bidDocumentFee', '-'))
add_row('第2类：财务信息', '投标保证金', financialInfo.get('bidBond', '-'))
add_row('第2类：财务信息', '履约保证金', financialInfo.get('performanceBond', '-'))
add_row('第2类：财务信息', '质量保证金', financialInfo.get('qualityBond', '-'))
add_row('第2类：财务信息', '代理费', financialInfo.get('agencyFee', '-'))

# ==================== 第3类：资质要求 ====================
qualReqs = data.get('qualificationRequirements', [])
for q in qualReqs:
    add_row('第3类：资质要求', '资质要求: ' + q.get('name', '-'), q.get('description', '-'), 
            '实质性要求' if q.get('isSubstantial') else '')
    if q.get('jointBid'):
        add_row('第3类：资质要求', '联合体投标', '允许（成员单位不超过两家）')
    if q.get('creditRequirements'):
        add_row('第3类：资质要求', '信用要求', q.get('creditRequirements'))

# ==================== 第4类：评分规则 ====================
scoringRules = data.get('scoringRules', {})
add_row('第4类：评分规则', '总分', f"{scoringRules.get('totalScore', 100)}分", '', '', '', '是')
add_row('第4类：评分规则', '技术分', f"{scoringRules.get('technicalScore', 0)}分", '', '', '', '是')
add_row('第4类：评分规则', '商务分', f"{scoringRules.get('commercialScore', 0)}分", '', '', '', '是')
add_row('第4类：评分规则', '价格分', f"{scoringRules.get('priceScore', 0)}分", '', '', '', '是')
add_row('第4类：评分规则', '中标方式', scoringRules.get('winningMethod', '-'))
add_row('第4类：评分规则', '评标方式', scoringRules.get('evaluationMethod', '-'))
if scoringRules.get('voidBidExplanation'):
    add_row('第4类：评分规则', '废标说明', scoringRules.get('voidBidExplanation'), '', '高', '是', '', '重点关注')

# 评分项明细
for item in scoringRules.get('scoringItems', []):
    add_row('第4类：评分规则', f"评分项: {item.get('name', '-')}", 
            f"最高{item.get('maxScore', 0)}分 | {item.get('description', '-')}",
            item.get('calculationMethod', ''), '', '', '是')

# 要求的证书
certs = scoringRules.get('requiredCompanyCertificates', [])
if certs:
    add_row('第4类：评分规则', '要求企业证书', '、'.join(certs), '', '高', '', '是', '确认是否具备')

# ==================== 第5类：时间要求 ====================
timeReqs = data.get('timeRequirements', {})
add_row('第5类：时间要求', '获取招标文件截止', timeReqs.get('documentAcquisitionDeadline', '-'))
add_row('第5类：时间要求', '标前提问截止', timeReqs.get('preBidQuestionDeadline', '-'))
add_row('第5类：时间要求', '开标时间', timeReqs.get('bidOpeningTime', '-'))
add_row('第5类：时间要求', '中标交货时间', timeReqs.get('winningDeliveryTime', '-'))
add_row('第5类：时间要求', '合同履约期限', timeReqs.get('contractPerformancePeriod', '-'))

# ==================== 第6类：项目信息 ====================
projectInfo = data.get('projectInfo', {})
add_row('第6类：项目信息', '实质性要求', projectInfo.get('substantialRequirements', '-'))
add_row('第6类：项目信息', '偏离结果', projectInfo.get('deviationResult', '-'))
add_row('第6类：项目信息', '现场踏勘', projectInfo.get('siteSurveyRequired', '-'))
add_row('第6类：项目信息', '控标点', projectInfo.get('controlPoints', '-'))
add_row('第6类：项目信息', '商务要求', projectInfo.get('businessRequirements', '-'))
add_row('第6类：项目信息', '技术需求', projectInfo.get('technicalRequirements', '-'))
add_row('第6类：项目信息', '核心服务要求', projectInfo.get('coreServiceRequirements', '-'))
add_row('第6类：项目信息', '密封要求', projectInfo.get('sealingRequirements', '-'))
add_row('第6类：项目信息', '验收要求', projectInfo.get('acceptanceRequirements', '-'))

# ==================== 第7类：电话问题（AI分析生成） ====================
phoneQuestions = data.get('phoneQuestions', [])
for idx, q in enumerate(phoneQuestions, 1):
    priority = q.get('priority', 'medium')
    riskLevel = '高' if priority == 'high' else ('中' if priority == 'medium' else '低')
    add_row('第7类：电话问题', f"问题{idx}: {q.get('question', '-')}", 
            q.get('reason', '-'), q.get('category', ''), riskLevel, '', '', '电话确认')

# ==================== 风险清单 ====================
risks = data.get('risks', [])
for risk in risks:
    level = risk.get('level', 'medium')
    levelText = {'critical': '严重', 'high': '高', 'medium': '中', 'low': '低'}.get(level, '中')
    add_row('风险清单', risk.get('title', '-'), risk.get('description', '-'),
            risk.get('source', ''), levelText, 
            '是' if risk.get('category') == 'void' else '否',
            '是' if risk.get('category') == 'score' else '否',
            risk.get('suggestion', ''))

# ==================== 准备清单 ====================
checklist = data.get('checklist', [])
for item in checklist:
    add_row('准备清单', item.get('item', '-'), item.get('category', '-'),
            item.get('source', ''), '', '', 
            f"{item.get('scoreWeight', 0)}分" if item.get('scoreWeight') else '',
            '必须准备' if item.get('required') else '建议准备')

# ==================== 投标建议 ====================
recommendation = data.get('recommendation', 'caution')
recText = {'bid': '建议投标', 'no-bid': '不建议投标'}.get(recommendation, '谨慎投标')
add_row('投标建议', '建议', recText, '；'.join(data.get('reasons', [])))

# 保存文件
out_path = r'C:\Users\ips\Desktop\测试AI员工能力\重做投标网站\杭州非遗项目-投标决策评估表.xlsx'
wb.save(out_path)

# 统计
print(f'导出成功: {out_path}')
print(f'\n统计:')
print(f'  总行数: {row - 2}')
print(f'  第1类：基本信息: 12行')
print(f'  第2类：财务信息: 10行')
print(f'  第3类：资质要求: 3行')
print(f'  第4类：评分规则: 8行')
print(f'  第5类：时间要求: 5行')
print(f'  第6类：项目信息: 9行')
print(f'  第7类：电话问题: {len(phoneQuestions)}行 (AI分析生成)')
print(f'  风险清单: {len(risks)}行')
print(f'  准备清单: {len(checklist)}行')
print(f'  投标建议: 1行')
