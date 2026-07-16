# -*- coding: utf-8 -*-
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\ips\Desktop\Upan\杭州非遗\投标决策杭州文化两中心.xlsx'

wb = load_workbook(file_path)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n{"="*60}')
    print(f'工作表: {sheet_name}')
    print(f'{"="*60}')
    
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=False):
        values = []
        for cell in row:
            val = cell.value if cell.value is not None else ''
            values.append(str(val))
        
        # 只打印有内容的行
        if any(v.strip() for v in values):
            print(' | '.join(values))
